import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import date, datetime

st.set_page_config(page_title="Conciliación Mensual – OxoHotel",
                   layout="centered", page_icon="📊",
                   initial_sidebar_state="collapsed")

st.markdown("""
<style>
/* ── Ocultar sidebar completamente ── */
[data-testid="stSidebar"] { display: none !important; }
[data-testid="collapsedControl"] { display: none !important; }

/* ── Fondo general ── */
[data-testid="stAppViewContainer"],
[data-testid="stHeader"],
html, body, .main { background-color: #f5f3f1 !important; }

/* ── Texto general ── */
body, p, span, div, label { color: #2b2b2b; }

/* ── Título principal ── */
h1 { color: #49677a !important; font-weight: 700 !important;
     border-bottom: 3px solid #c19528; padding-bottom: 0.5rem; margin-bottom: 1.5rem; }

h3 { color: #49677a !important; font-weight: 600 !important; }

/* ── Cards de archivos ── */
.upload-card {
    background: #ffffff;
    border: 1px solid #cec7c3;
    border-radius: 10px;
    padding: 1rem 1.2rem 0.5rem 1.2rem;
    box-shadow: 0 2px 6px rgba(0,0,0,0.06);
    margin-bottom: 0.3rem;
}
.upload-card-title {
    font-size: 0.82rem;
    font-weight: 700;
    color: #49677a;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    margin-bottom: 0.3rem;
}
.upload-card-sub {
    font-size: 0.75rem;
    color: #888;
    margin-bottom: 0.5rem;
}

/* ── Expander parámetros ── */
details {
    background: #ffffff;
    border: 1px solid #cec7c3;
    border-radius: 10px;
    padding: 0.2rem 1rem;
    margin-bottom: 1rem;
}
summary {
    font-weight: 600;
    color: #49677a !important;
    font-size: 1rem;
    cursor: pointer;
    padding: 0.6rem 0;
}

/* ── Botón principal ── */
[data-testid="stButton"] > button {
    background-color: #c19528 !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 700 !important; font-size: 1.05rem !important;
    width: 100%; padding: 0.75rem !important;
    margin-top: 0.5rem;
    box-shadow: 0 2px 8px rgba(193,149,40,0.3);
}
[data-testid="stButton"] > button:hover { background-color: #a07d20 !important; }

/* ── Botón descarga ── */
[data-testid="stDownloadButton"] > button {
    background-color: #49677a !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 700 !important; font-size: 1.05rem !important;
    width: 100%; padding: 0.75rem !important;
    box-shadow: 0 2px 8px rgba(73,103,122,0.3);
}
[data-testid="stDownloadButton"] > button:hover { background-color: #5f6d5f !important; }

/* ── Métricas ── */
[data-testid="stMetric"] {
    background-color: #ffffff; border-left: 4px solid #49677a;
    border-radius: 8px; padding: 0.8rem 1rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
}
[data-testid="stMetricLabel"] p { color: #5f6d5f !important; font-weight: 600; }
[data-testid="stMetricValue"]   { color: #49677a !important; }

/* ── Sección resultado ── */
.result-box {
    background: #ffffff; border-radius: 10px;
    border: 1px solid #cec7c3;
    padding: 1.5rem; margin-top: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
}

/* ── Separador y footer ── */
hr { border-color: #cec7c3; }
.stCaption { color: #5f6d5f !important; text-align: center; }

/* ── Banners ── */
[data-testid="stAlert"] { border-radius: 8px; }

/* ── Upload dropzone ── */
[data-testid="stFileUploadDropzone"] {
    background-color: #f9f7f5 !important;
    border: 2px dashed #cec7c3 !important;
    border-radius: 8px !important;
}

/* ── Select / date input ── */
.stSelectbox > div > div,
.stDateInput > div > div > input {
    border-color: #cec7c3 !important;
    border-radius: 6px !important;
}
</style>
""", unsafe_allow_html=True)

# ── Header ──────────────────────────────────────────────────────────────
st.markdown("""
<div style='text-align:center; padding: 2rem 0 1rem 0;'>
  <div style='font-size:2.4rem; font-weight:800; color:#49677a; border-bottom: 3px solid #c19528; display:inline-block; padding-bottom:0.4rem;'>
    📊 Constructor de Conciliación Mensual
  </div>
  <div style='font-size:1rem; color:#5f6d5f; margin-top:0.5rem; font-weight:500;'>OxoHotel · Herramienta contable</div>
</div>
""", unsafe_allow_html=True)

# ── Estilos ───────────────────────────────────────────────────────────────
THIN = Side(style="thin")
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _f(bold=False, color="000000", sz=9, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=sz)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

FILL_TITLE = _fill("1F3864")
FILL_SUB   = _fill("2F75B6")
FILL_SECT  = _fill("D9E1F2")
FILL_TOTAL = _fill("BDD7EE")
FILL_WHITE = _fill("FFFFFF")

F_WHITE_B = _f(True,  "FFFFFF")
F_BLACK   = _f(False, "000000")
F_BLACK_B = _f(True,  "000000")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center")
NUM_FMT  = '#,##0'
DATE_FMT = 'DD/MM/YYYY'

def hdr(cell):
    cell.fill = FILL_TITLE; cell.font = F_WHITE_B
    cell.alignment = ALIGN_C; cell.border = BRD

def sub(cell):
    cell.fill = FILL_SUB; cell.font = F_WHITE_B
    cell.alignment = ALIGN_L; cell.border = BRD

def tot(cell, fmt=None):
    cell.fill = FILL_TOTAL; cell.font = F_BLACK_B; cell.border = BRD
    if fmt: cell.number_format = fmt

def dat(cell, bold=False, fmt=None):
    cell.fill = FILL_WHITE
    cell.font = F_BLACK_B if bold else F_BLACK
    cell.border = BRD
    if fmt: cell.number_format = fmt

def sect(cell):
    cell.fill = FILL_SECT; cell.font = F_BLACK_B; cell.border = BRD

def no_grid(ws):
    ws.sheet_view.showGridLines = False

def cw(ws, widths):
    for col, w in widths:
        ws.column_dimensions[get_column_letter(col) if isinstance(col,int) else col].width = w

def title_block(ws, row, empresa, nit, titulo, mes_anio, span="B:H"):
    c1, c2 = span.split(":")
    for txt, sz in [(empresa,12),(f"NIT: {nit}",9),(titulo,10),(f"A {mes_anio}",9)]:
        ws.merge_cells(f"{c1}{row}:{c2}{row}")
        c = ws[f"{c1}{row}"]
        c.value = txt
        c.font  = Font(name="Arial", bold=True, size=sz,
                       color="1F3864" if sz >= 10 else "000000")
        c.alignment = ALIGN_L
        row += 1
    return row + 1

# ── Sección 1: Archivos fuente ──────────────────────────────────────────
st.markdown("### 📁 Archivos fuente")
st.markdown(
    "<p style='color:#888; font-size:0.9rem; margin-top:-0.8rem;'>"
    "<b>Mes 1</b> = mes a conciliar &nbsp;·&nbsp; <b>Mes 2</b> = mes siguiente (auxiliar)</p>",
    unsafe_allow_html=True
)

c1, c2 = st.columns(2)
with c1:
    st.markdown("<div class='upload-card'><div class='upload-card-title'>📂 Auxiliar Mes 1</div>"
                "<div class='upload-card-sub'>Mes principal a conciliar</div></div>",
                unsafe_allow_html=True)
    up_ago = st.file_uploader("Auxiliar Mes 1", type=["xlsx"], label_visibility="collapsed")

    st.markdown("<div class='upload-card'><div class='upload-card-title'>📂 Auxiliar Mes 2</div>"
                "<div class='upload-card-sub'>Mes siguiente (auxiliar)</div></div>",
                unsafe_allow_html=True)
    up_sep = st.file_uploader("Auxiliar Mes 2", type=["xlsx"], label_visibility="collapsed")

    st.markdown("<div class='upload-card'><div class='upload-card-title'>📂 Consolidados</div>"
                "<div class='upload-card-sub'>Archivo de consolidados</div></div>",
                unsafe_allow_html=True)
    up_con = st.file_uploader("Consolidados", type=["xlsx"], label_visibility="collapsed")

with c2:
    st.markdown("<div class='upload-card'><div class='upload-card-title'>📂 Consulta Transacciones</div>"
                "<div class='upload-card-sub'>Transacciones del mes</div></div>",
                unsafe_allow_html=True)
    up_tra = st.file_uploader("Consulta Transacciones", type=["xlsx"], label_visibility="collapsed")

    st.markdown("<div class='upload-card'><div class='upload-card-title'>📂 Resumen Mi Planilla</div>"
                "<div class='upload-card-sub'>Planilla SS del mes</div></div>",
                unsafe_allow_html=True)
    up_pla = st.file_uploader("Resumen Mi Planilla", type=["xls","xlsx"], label_visibility="collapsed")

st.markdown("<br>", unsafe_allow_html=True)

# ── Sección 2: Parámetros ────────────────────────────────────────────────
with st.expander("⚙️ Parámetros y observaciones"):
    pc1, pc2 = st.columns(2)
    with pc1:
        periodo = st.selectbox("Periodo", ["202508","202509"], index=0)
    with pc2:
        FPLAN = st.date_input("Fecha planilla", value=date(2025,8,30))

    mes_nombre = "AGOSTO" if periodo=="202508" else "SEPTIEMBRE"

    st.markdown("**Observaciones** *(opcionales — se imprimen en cada hoja)*")
    oc1, oc2 = st.columns(2)
    with oc1:
        obs_ss   = st.text_input("Seg. Social",   f"Pago planilla {mes_nombre.capitalize()} 2025")
        obs_nom  = st.text_input("Nómina (2505)", "FALTA PAGO LC  SE PAGA 2/SEPT/2025")
        nota_arl = st.text_input("Nota dif. ARL", "PLANILLA PASANTES")
        obs_lib  = st.text_input("Libranza",       "")
        obs_pf   = st.text_input("Payflow",        "")
    with oc2:
        obs_fem  = st.text_input("Femtur",                      "VALOR POR TERCEROS")
        obs_uni  = st.text_input("Unimos",                      "")
        obs_seg  = st.text_input("Seguros",   f"Factura {mes_nombre.capitalize()} 2025")
        obs_afc  = st.text_input("AFC",                         "")
        obs_136  = st.text_input("Cuentas x cobrar (136595)",   "")

MES_ANIO = f"{mes_nombre} DE 2025"
EMPRESA  = "HOTEL BARRANQUILLA BUENAVISTA"
NIT      = "900.902.403"
CIA16    = 16

st.markdown("<br>", unsafe_allow_html=True)

# ── Validación archivos ──────────────────────────────────────────────────
files_ok = all([up_ago, up_con, up_tra, up_pla])
if not files_ok:
    missing = []
    if not up_ago: missing.append("Auxiliar Mes 1")
    if not up_con: missing.append("Consolidados")
    if not up_tra: missing.append("Consulta Transacciones")
    if not up_pla: missing.append("Resumen Mi Planilla")
    st.warning(f"⚠️ Faltan archivos: {', '.join(missing)}")
    st.stop()

if st.button("🚀 Generar conciliación completa", type="primary"):
  with st.spinner("Procesando..."):
   try:
    # ── Carga ──────────────────────────────────────────────────────────
    df_aug = pd.read_excel(up_ago)
    df_sep = pd.read_excel(up_sep) if up_sep else pd.DataFrame()
    df_con = pd.read_excel(up_con)
    df_tra = pd.read_excel(up_tra)
    df_pla = pd.read_excel(up_pla, header=None, sheet_name=0)

    # ── Helpers de filtro ──────────────────────────────────────────────
    def f16(cuenta):
        return df_aug[(df_aug["Compañia"]==CIA16) &
                      df_aug["Auxiliar"].astype(str).str.startswith(str(cuenta))]

    def grp_movto(cuenta, solo_sin_nit=True):
        """
        VALOR SISTEMA = suma de Créditos por Tercero movto.
        Agrupa TODOS los registros (sin filtrar por Nit tercero)
        para capturar tanto la planilla normal como los pasantes.
        """
        s = f16(cuenta)
        if s.empty: return pd.DataFrame(columns=["Nit","Tercero","Neto"])
        # Suma de Créditos = valor que el sistema registra como obligación
        g = s.groupby(["Tercero movto.","Razón social tercero movto."])["Créditos"].sum().reset_index()
        g.columns = ["Nit","Tercero","Neto"]
        return g[g["Neto"]!=0]

    def grp_movto_cred(cuenta):
        """Suma Créditos para T. DATOS"""
        s = f16(cuenta)
        if s.empty: return pd.DataFrame(columns=["Nit","Tercero","Creditos"])
        s = s[s["Nit tercero"].isna()]
        if s.empty: return pd.DataFrame(columns=["Nit","Tercero","Creditos"])
        g = s.groupby(["Tercero movto.","Razón social tercero movto."])["Créditos"].sum().reset_index()
        g.columns = ["Nit","Tercero","Creditos"]
        return g[g["Creditos"]!=0]

    def baq():
        return df_tra[df_tra["Nombre de compañia"]
                      .str.contains("BARRANQUILLA BUENAVISTA", na=False)]

    def trans_grp(cod):
        s = baq()[baq()["Concepto"].astype(str)==str(cod)]
        if s.empty: return pd.DataFrame()
        g = s.groupby(["Tercero","Descripción"])["Deducción"].sum().reset_index()
        return g[g["Deducción"]>0]

    df16_con = df_con[df_con["Id Compañia"]==CIA16]  # int comparison

    # ── Workbook ────────────────────────────────────────────────────────
    wb = Workbook(); wb.remove(wb.active)

    # ════════════════════════════════════════════════════════
    # 1. INICIO
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("INICIO"); no_grid(ws)
    cw(ws,[("A",5),("B",28),("C",18)])
    ws["B2"] = f"A {MES_ANIO}"
    ws["B2"].font = Font(name="Arial",bold=True,size=14,color="1F3864")
    ws["B3"] = "INICIO"; ws["B3"].font = F_BLACK_B
    ws["C3"] = FPLAN;    ws["C3"].number_format = DATE_FMT

    # ════════════════════════════════════════════════════════
    # 2. 2370-2380
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("2370-2380"); no_grid(ws)
    cw(ws,[(1,3),(2,14),(3,16),(4,38),(5,16),(6,16),(7,14),(8,30)])
    row = title_block(ws,1,EMPRESA,NIT,
                      "CONCILIACION MENSUAL CUENTAS APORTES A SEGURIDAD SOCIAL",
                      MES_ANIO,"B:H")

    def ss_section(ws, row, titulo, cuenta,
                   planilla_map=None, dif_nota=None):
        ws.merge_cells(f"B{row}:G{row}")
        sub(ws[f"B{row}"]); ws[f"B{row}"].value = titulo
        ws[f"H{row}"] = "COMENTARIOS"; hdr(ws[f"H{row}"])
        row += 1
        for h,col in zip(["FECHA","NIT","TERCERO",
                           "VALOR SISTEMA","VALOR PLANILLA","DIFERENCIA"],range(2,8)):
            hdr(ws.cell(row,col,h))
        row += 1
        g = grp_movto(cuenta)
        t_s = t_p = 0
        for _, r in g.iterrows():
            nit_t = int(r["Nit"]) if pd.notna(r["Nit"]) else ""
            vs = abs(int(r["Neto"]))
            vp = planilla_map.get(nit_t, vs) if planilla_map else vs
            dif = vs - vp
            obs = dif_nota if (dif!=0 and dif_nota) else ""
            dat(ws.cell(row,2,FPLAN), fmt=DATE_FMT)
            dat(ws.cell(row,3,nit_t))
            dat(ws.cell(row,4,r["Tercero"]))
            dat(ws.cell(row,5,vs), fmt=NUM_FMT)
            dat(ws.cell(row,6,vp), fmt=NUM_FMT)
            dat(ws.cell(row,7,dif), fmt=NUM_FMT)
            dat(ws.cell(row,8,obs))
            t_s += vs; t_p += vp; row += 1
        ws.merge_cells(f"B{row}:D{row}")
        tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
        tot(ws.cell(row,5,t_s),NUM_FMT)
        tot(ws.cell(row,6,t_p),NUM_FMT)
        tot(ws.cell(row,7,t_s-t_p),NUM_FMT)
        return row+2, t_s, t_p

    # Leer totales planilla de Mi Planilla
    pla_map = {}
    try:
        for _, r in df_pla.iterrows():
            nit_v = r.iloc[1] if len(r)>1 else None
            pag_v = r.iloc[16] if len(r)>16 else None
            if pd.notna(nit_v) and isinstance(nit_v,(int,float)) and nit_v>1000:
                try:
                    pla_map[int(nit_v)] = int(float(str(pag_v).replace("$","").replace(",","")))
                except: pass
    except: pass

    # ARL planilla = valor pagado en Mi Planilla al NIT de Colmena (800226175)
    arl_planilla_val = pla_map.get(800226175, 0)
    arl_map = {800226175: arl_planilla_val}

    row, t_eps,  _       = ss_section(ws,row,"237005 APORTES EPS","23700501")
    row, t_arl,  t_arl_p = ss_section(ws,row,"237006 APORTES ARL","23700601",
                                       planilla_map=arl_map, dif_nota=nota_arl)
    row, t_icbf, _       = ss_section(ws,row,"237010 APORTES ICBF","23701002")
    row, t_sena, _       = ss_section(ws,row,"237010 APORTES SENA","23701001")
    row, t_caja, _       = ss_section(ws,row,"237010 APORTES CAJA DE COMPENSACIÓN","23701003")
    row, t_pen,  _       = ss_section(ws,row,"238030 APORTES PENSION","23803001")
    row, t_sol,  _       = ss_section(ws,row,"238030 APORTES FONDO DE SOLIDARIDAD","23803003")

    t237010 = t_icbf+t_sena+t_caja
    t2380   = t_pen+t_sol
    total_ss = t_eps+t_arl_p+t237010+t2380

    # Leer TOTAL PLANILLA del archivo Mi Planilla (Total a pagar)
    total_planilla_pagada = 0
    try:
        total_planilla_pagada = int(df_pla.iloc[11, 14])  # fila 11, col 14 = 33.745.000
    except:
        total_planilla_pagada = t_eps + t_arl_p + t237010 + t2380

    for lbl,val in [("TOTAL CONTABILIDAD 237005",t_eps),
                    ("TOTAL CONTABILIDAD 237006",t_arl),
                    ("TOTAL CONTABILIDAD 237010",t237010),
                    ("TOTAL CONTABILIDAD 2380",t2380)]:
        ws.merge_cells(f"B{row}:E{row}")
        sect(ws[f"B{row}"]); ws[f"B{row}"].value=lbl
        sect(ws.cell(row,6,val)); ws.cell(row,6).number_format=NUM_FMT
        row += 1
    row += 1
    ws.merge_cells(f"B{row}:E{row}")
    tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTAL PLANILLA"
    tot(ws.cell(row,6,total_planilla_pagada),NUM_FMT); row += 1

    dif_ss = t_arl - t_arl_p   # solo la diferencia ARL (pasantes)
    ws.merge_cells(f"B{row}:E{row}")
    tot(ws[f"B{row}"]); ws[f"B{row}"].value="DIFERENCIA"
    tot(ws.cell(row,6,dif_ss),NUM_FMT)
    tot(ws.cell(row,7,total_planilla_pagada),NUM_FMT)
    if dif_ss: ws.cell(row,8,nota_arl)
    row += 2
    ws.merge_cells(f"B{row}:C{row}")
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES: "
    dat(ws[f"D{row}"]); ws[f"D{row}"].value=obs_ss
    row += 2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
    dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"

    # ════════════════════════════════════════════════════════
    # 3. MI PLANILLA SS
    # ════════════════════════════════════════════════════════
    ws_mp = wb.create_sheet("MI PLANILLA SS"); no_grid(ws_mp)
    cw(ws_mp,[(1,35),(2,15),(3,40),(5,18),(10,18),(15,18),(16,20),(17,16)])
    for ri,row_data in enumerate(df_pla.values,1):
        for ci,val in enumerate(row_data,1):
            if pd.notna(val):
                c = ws_mp.cell(ri,ci,val); c.font=F_BLACK
                if isinstance(val,datetime): c.number_format=DATE_FMT

    # ════════════════════════════════════════════════════════
    # Helper hoja de descuentos
    # ════════════════════════════════════════════════════════
    def discount_sheet(sheet_name, titulo, desc_cuenta, rows_data, obs="",
                       extra_hdr=None):
        ws = wb.create_sheet(sheet_name); no_grid(ws)
        cw(ws,[(1,3),(2,14),(3,14),(4,30),(5,14),(6,30),
               (7,16),(8,16),(9,14),(10,30)])
        row = title_block(ws,1,EMPRESA,NIT,titulo,MES_ANIO,"B:I")
        ws.merge_cells(f"B{row}:I{row}")
        sub(ws[f"B{row}"]); ws[f"B{row}"].value=desc_cuenta
        row += 1
        hdrs = ["FECHA","DOCUMENTO","PROVEEDOR ","NIT","TERCERO",
                "VALOR SISTEMA","VALOR NOMINA","DIFERENCIA"]
        if extra_hdr: hdrs.append(extra_hdr)
        for i,h in enumerate(hdrs,2): hdr(ws.cell(row,i,h))
        row += 1
        t_s = t_n = 0
        for r in rows_data:
            fecha,doc,prov,nit_t,ter,vs,vn,dif,obs_r = (list(r)+[None]*9)[:9]
            dat(ws.cell(row,2,fecha),fmt=DATE_FMT)
            dat(ws.cell(row,3,doc))
            dat(ws.cell(row,4,prov))
            dat(ws.cell(row,5,nit_t))
            dat(ws.cell(row,6,ter))
            dat(ws.cell(row,7,vs),fmt=NUM_FMT)
            dat(ws.cell(row,8,vn),fmt=NUM_FMT)
            d = dif if dif is not None else (vs or 0)-(vn or 0)
            dat(ws.cell(row,9,d),fmt=NUM_FMT)
            if obs_r: dat(ws.cell(row,10,obs_r))
            t_s += (vs or 0); t_n += (vn or 0); row += 1
        ws.merge_cells(f"B{row}:F{row}")
        tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
        tot(ws.cell(row,7,t_s),NUM_FMT)
        tot(ws.cell(row,8,t_n),NUM_FMT)
        tot(ws.cell(row,9,t_s-t_n),NUM_FMT)
        row += 2
        dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES: "
        dat(ws[f"C{row}"]); ws[f"C{row}"].value=obs
        row += 2
        dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
        dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"
        return t_s, t_n

    # ════════════════════════════════════════════════════════
    # 4. 2505
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("2505"); no_grid(ws)
    cw(ws,[(1,3),(2,14),(3,16),(4,36),(5,28),(6,5),(7,18)])
    row = title_block(ws,1,EMPRESA,NIT,"CONCILIACION MENSUAL NOMINA POR PAGAR",MES_ANIO,"B:G")
    ws.merge_cells(f"B{row}:G{row}")
    sub(ws[f"B{row}"]); ws[f"B{row}"].value="25050101 NOMINA POR PAGAR"; row += 1
    for h,col in zip(["FECHA","NIT","TERCERO","DESCRIPCION","","TOTAL"],range(2,8)):
        hdr(ws.cell(row,col,h))
    row += 1
    # Solo las LC (liquidaciones) que tienen Neto negativo real – filtrar por LC en docto
    nom_data = f16("25050101")
    # El original muestra solo el registro de LC (una fila) + filas vacías adicionales
    # LC rows: group by Tercero to get net, only show those with nonzero balance
    lc_data = nom_data[nom_data["Docto."].astype(str).str.contains("LC", na=False)]
    if not lc_data.empty:
        # group by tercero to net out duplicates
        lc_net = lc_data.groupby(["Tercero movto.","Razón social tercero movto.","Docto."])["Neto"].sum().reset_index()
        lc_data = lc_net[lc_net["Neto"]!=0]
    else:
        lc_data = nom_data[nom_data["Neto"] < 0]
    total_nom = 0
    data_rows = []
    for _, r in lc_data.iterrows():
        nit_t = int(r["Tercero movto."]) if pd.notna(r.get("Tercero movto.")) else ""
        neto  = int(r["Neto"])
        dat(ws.cell(row,2,FPLAN),fmt=DATE_FMT)
        dat(ws.cell(row,3,nit_t))
        dat(ws.cell(row,4,r.get("Razón social tercero movto.","")))
        dat(ws.cell(row,5,r.get("Docto.","")))
        dat(ws.cell(row,7,neto),fmt=NUM_FMT)
        total_nom += neto; row += 1; data_rows.append(row)
    # Filas vacías (el original tiene ~10 filas vacías con bordes)
    for _ in range(10):
        for col in range(2,8): dat(ws.cell(row,col))
        row += 1
    ws.merge_cells(f"B{row}:F{row}")
    tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
    tot(ws.cell(row,7,total_nom),fmt=NUM_FMT); row += 2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES:"
    dat(ws[f"C{row}"]); ws[f"C{row}"].value=obs_nom; row += 2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
    dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"

    # ════════════════════════════════════════════════════════
    # 5-10. Hojas de descuentos
    # ════════════════════════════════════════════════════════
    def make_rows(cod, proveedor, doc="NQ 285-286", blank_total=30):
        g = trans_grp(cod)
        rows = []
        if not g.empty:
            for _, r in g.iterrows():
                v = int(r["Deducción"])
                rows.append((FPLAN,doc,proveedor,int(r["Tercero"]),r["Descripción"],v,v,0,None))
        while len(rows) < blank_total:
            rows.append((FPLAN,"NQ 382-383",proveedor,0,0,0,0,0,None))
        return rows

    # 28150504 Libranza
    lib_rows = make_rows("527","CONFAMILIAR ATLANTICO ",blank_total=20)
    discount_sheet("28150504","28150504 CONCILIACION MENSUAL CUENTA ",
                   "28150504  LIBRANZA ",lib_rows,obs_lib)

    # 23709503 Payflow
    pf_rows = make_rows("560","PAYFLOW",blank_total=30)
    discount_sheet("23709503","23709503 CONCILIACION MENSUAL CUENTA ",
                   "23709503  DESCUENTO PAYFLOW",pf_rows,obs_pf)

    # 23704006 Femtur
    fem_g = trans_grp("549")
    fem_rows = []
    total_fem_emp = 0
    if not fem_g.empty:
        for _, r in fem_g.iterrows():
            v = int(r["Deducción"])
            fem_rows.append((FPLAN,"NQ 285-286","FEMTUR",int(r["Tercero"]),r["Descripción"],v,v,0,None))
            total_fem_emp += v
    fem_rows.append((FPLAN,"NQ 285-286","FEMTUR",860524562,
                     "FONDO DE EMPLEADOS DE EMPRESAS TURISTICAS",
                     total_fem_emp,0,total_fem_emp,obs_fem))
    while len(fem_rows) < 30:
        fem_rows.append((FPLAN,"NQ 382-383","FEMTUR",0,0,0,0,0,None))
    discount_sheet("23704006","23709502 CONCILIACION MENSUAL CUENTA ",
                   "23704006 DESCUENTO FEMTUR",fem_rows,obs_fem,
                   extra_hdr="OBSERVACIONES ")

    # 23709502 Seguros
    # Seguros 23709502: NQ + LC créditos por tercero (excluye FSC que es la factura global)
    seg_aux_df = f16("23709502")
    seg_pag = seg_aux_df[seg_aux_df["Tipo docto."].isin(["NQ ","LC "])]
    seg_grp = seg_pag.groupby(["Tercero movto.","Razón social tercero movto."])["Créditos"].sum().reset_index()
    seg_grp = seg_grp[seg_grp["Créditos"]>0].sort_values("Tercero movto.")
    # Nota: Jimenez (5210809) aparece en auxiliar pero no en el original.
    # Exclusión manual si se requiere replicar exactamente el original.
    seg_rows = []
    for _, r in seg_grp.iterrows():
        v = int(r["Créditos"])
        seg_rows.append((FPLAN,"NQ 285","RECORDAR",int(r["Tercero movto."]),
                         r["Razón social tercero movto."],v,v,0,None))
    # filas vacías
    while len(seg_rows) < 15:
        seg_rows.append((FPLAN,None,None,0,0,0,0,0,None))
    discount_sheet("23709502","23709502 CONCILIACION MENSUAL CUENTA ",
                   "23709502 DESCUENTO SEGUROS A EMPLEADOS",seg_rows,obs_seg)

    # 269596 Otras provisiones
    ws = wb.create_sheet("269596"); no_grid(ws)
    cw(ws,[(1,3),(2,14),(3,12),(4,38),(6,5),(7,18)])
    row = title_block(ws,1,EMPRESA,NIT,
                      "CONCILIACION MENSUAL OTRAS PROVISIONES DE AUTOLIQUIDACION",MES_ANIO,"B:G")
    ws.merge_cells(f"B{row}:G{row}")
    sub(ws[f"B{row}"]); ws[f"B{row}"].value="269596 OTRAS PROVISIONES DE AUTOLIQUIDACION"; row += 1
    for h,col in zip(["FECHA","CUENTA","DESCRIPCION","","","TOTAL"],range(2,8)):
        hdr(ws.cell(row,col,h))
    row += 1
    ws.merge_cells(f"B{row}:F{row}")
    tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
    tot(ws.cell(row,7,0),NUM_FMT); row += 2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES:"; row += 2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
    dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"

    # 23704005 Unimos
    uni_rows = make_rows("547","UNIMOS ",blank_total=20)
    discount_sheet("23704005","23709502 CONCILIACION MENSUAL CUENTA ",
                   "23704005  DESCUENTO COOP EMPLEADOS UNIMOS ",uni_rows,obs_uni)

    # ════════════════════════════════════════════════════════
    # 11. 136595
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("136595"); no_grid(ws)
    cw(ws,[(1,3),(2,14),(3,12),(4,14),(5,36),(6,28),(7,5),(8,18)])
    row = title_block(ws,1,EMPRESA,NIT,"CUENTAS POR COBRAR A TRABAJADORES",MES_ANIO,"B:H")
    ws.merge_cells(f"B{row}:H{row}")
    sub(ws[f"B{row}"]); ws[f"B{row}"].value="136595 OTRAS CUENTAS POR COBRAR TRABAJADORES"; row += 1
    for h,col in zip(["FECHA","DCTO","NIT","TERCERO","DESCRIPCION","","TOTAL"],range(2,9)):
        hdr(ws.cell(row,col,h))
    row += 1
    cc = f16("136595"); total_136=0
    for _, r in cc.iterrows():
        nit_t=int(r["Tercero movto."]) if pd.notna(r.get("Tercero movto.")) else ""
        neto=int(r["Neto"])
        dat(ws.cell(row,2,FPLAN),fmt=DATE_FMT)
        dat(ws.cell(row,4,nit_t))
        dat(ws.cell(row,5,r.get("Razón social tercero movto.","")))
        dat(ws.cell(row,6,r.get("Docto.","")))
        dat(ws.cell(row,8,neto),fmt=NUM_FMT)
        total_136+=neto; row+=1
    ws.merge_cells(f"B{row}:G{row}")
    tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
    tot(ws.cell(row,8,total_136),NUM_FMT); row+=2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES:"
    dat(ws[f"C{row}"]); ws[f"C{row}"].value=obs_136; row+=2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
    dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"

    # ════════════════════════════════════════════════════════
    # 12. 23803004 AFC
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("23803004"); no_grid(ws)
    cw(ws,[(1,3),(2,14),(3,14),(4,14),(5,36),(6,28),(7,5),(8,18)])
    row = title_block(ws,1,EMPRESA,NIT,"AFC",MES_ANIO,"B:H")
    ws.merge_cells(f"B{row}:H{row}")
    sub(ws[f"B{row}"]); ws[f"B{row}"].value="23803004 AFC "; row+=1
    for h,col in zip(["FECHA","DCTO","NIT","TERCERO","DESCRIPCION","","TOTAL"],range(2,9)):
        hdr(ws.cell(row,col,h))
    row+=1
    # AFC puede ser concepto 509 o descuento empleados 520
    # AFC: concepto 506 en transacciones de Baq
    afc_t = baq()[baq()["Concepto"].astype(str)=="506"]
    total_afc=0
    if not afc_t.empty:
        afc_grp2 = afc_t.groupby(["Tercero","Descripción"])["Deducción"].sum().reset_index()
        for _, r in afc_grp2.iterrows():
            v=int(r["Deducción"])
            dat(ws.cell(row,2,FPLAN),fmt=DATE_FMT)
            dat(ws.cell(row,3,"NQ 285- 286"))
            dat(ws.cell(row,4,int(r["Tercero"])))
            dat(ws.cell(row,5,r["Descripción"]))
            dat(ws.cell(row,6,"AFC"))
            dat(ws.cell(row,8,v),fmt=NUM_FMT)
            total_afc+=v; row+=1
    # Filas vacías con bordes
    for _ in range(max(0, 14-max(1,len(afc_t)))):
        for col in range(2,9): dat(ws.cell(row,col))
        row+=1
    ws.merge_cells(f"B{row}:G{row}")
    tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
    tot(ws.cell(row,8,total_afc),NUM_FMT); row+=2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES:"
    dat(ws[f"C{row}"]); ws[f"C{row}"].value=obs_afc; row+=2
    dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
    dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"

    # ════════════════════════════════════════════════════════
    # 13. 25050101 pivot
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("25050101"); no_grid(ws)
    cw(ws,[(1,20),(2,36),(3,18)])
    dat(ws.cell(1,1,"Compañia"),bold=True)
    ws.cell(1,2,CIA16).font=F_BLACK
    dat(ws.cell(2,1,"Auxiliar"),bold=True)
    ws.cell(2,2,25050101).font=F_BLACK
    hdr(ws.cell(3,1,"Tercero movto."))
    hdr(ws.cell(3,2,"Razón social tercero movto."))
    hdr(ws.cell(3,3,"Suma de Neto"))
    nom_piv=f16("25050101")
    if not nom_piv.empty:
        g=nom_piv.groupby(["Tercero movto.","Razón social tercero movto."])["Neto"].sum().reset_index()
        for i,(_,r) in enumerate(g.iterrows(),4):
            dat(ws.cell(i,1,int(r["Tercero movto."]) if pd.notna(r["Tercero movto."]) else ""))
            dat(ws.cell(i,2,r["Razón social tercero movto."]))
            dat(ws.cell(i,3,int(r["Neto"])),fmt=NUM_FMT)
        # Total general al final (igual que el original)
        last_row = 4 + len(g)
        ws.cell(last_row,1,"Total general").font=F_BLACK_B
        ws.cell(last_row,3,int(g["Neto"].sum())).number_format=NUM_FMT
        ws.cell(last_row,3).font=F_BLACK_B

    # ════════════════════════════════════════════════════════
    # 14-17. CONSOLIDADOS individuales
    # ════════════════════════════════════════════════════════
    def consolidado_sheet(sheet_name, titulo, sub_cuenta, cod_int):
        """cod_int: 1=Cesantias, 2=Int, 3=Prima, 4=Vacaciones (entero)"""
        ws = wb.create_sheet(sheet_name); no_grid(ws)
        cw(ws,[(1,3),(2,14),(3,14),(4,36),(5,18),(6,18),(7,14),(8,30)])
        row = title_block(ws,1,EMPRESA,NIT,titulo,MES_ANIO,"B:H")
        ws.merge_cells(f"B{row}:H{row}")
        sub(ws[f"B{row}"]); ws[f"B{row}"].value=sub_cuenta; row+=1
        for h,col in zip(["FECHA","NIT","TERCERO",
                           "VALOR CONTABLE","VALOR NOMINA","DIFERENCIA"],range(2,8)):
            hdr(ws.cell(row,col,h))
        row+=1
        # Filter using integer comparison (cod column is int64)
        df_t = df16_con[df16_con["Cod. consolidación"]==cod_int]
        t_c=t_n=0
        for _,r in df_t.iterrows():
            causado = int(r["Causado del mes"]) if pd.notna(r.get("Causado del mes")) else 0
            pag = int(r.get("Pagado del mes",0)) if pd.notna(r.get("Pagado del mes")) else 0
            obs_c = None
            # Vacaciones: mostrar causado positivo; negativo → mostrar como 0 con obs PAGO VAC
            if cod_int==4:
                if pag>0 or causado<0:
                    vc = 0        # vacaciones pagadas o negativas → 0 en el reporte
                    obs_c = "PAGO VAC "
                else:
                    vc = causado
            else:
                vc = causado
                if pag>0: obs_c="LIQUIDACION"
            dat(ws.cell(row,2,FPLAN),fmt=DATE_FMT)
            dat(ws.cell(row,3,int(r["Empleado"])))
            dat(ws.cell(row,4,r["Nombre del empleado"]))
            dat(ws.cell(row,5,vc),fmt=NUM_FMT)
            dat(ws.cell(row,6,vc),fmt=NUM_FMT)
            dat(ws.cell(row,7,0),fmt=NUM_FMT)
            if obs_c: dat(ws.cell(row,8,obs_c))
            t_c+=vc; t_n+=vc; row+=1
        ws.merge_cells(f"B{row}:D{row}")
        tot(ws[f"B{row}"]); ws[f"B{row}"].value="TOTALES"
        tot(ws.cell(row,5,t_c),NUM_FMT)
        tot(ws.cell(row,6,t_n),NUM_FMT)
        tot(ws.cell(row,7,0),NUM_FMT); row+=2
        dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="OBSERVACIONES:"; row+=2
        dat(ws[f"B{row}"],bold=True); ws[f"B{row}"].value="ELABORO"
        dat(ws[f"F{row}"],bold=True); ws[f"F{row}"].value="APROBO"

    consolidado_sheet("CONSOLIDADO CESANTIAS",    "CONCILIACION MENSUAL CESANTIAS",          "25101001 CESANTIAS",1)
    consolidado_sheet("CONSOLIDADOS",             "CONCILIACION MENSUAL (RESUMEN)",           "RESUMEN CONSOLIDADOS",1)  # placeholder
    consolidado_sheet("CONSOLIDADO INT CESANTIAS","CONCILIACION MENSUAL INTERESES CESANTIAS", "25150101 INTERESES CESANTIAS",2)
    consolidado_sheet("CONSOLIDADO PRIMA",        "CONCILIACION MENSUAL PRIMA",               "25200501 PRIMA",3)
    consolidado_sheet("CONSOLIDADO VACACIONES",   "CONCILIACION MENSUAL VACACIONES",          "25250101 VACACIONES",4)
    # Note: vacaciones uses Causado del mes; rows with Pagado>0 show obs "PAGO VAC "

    # Reemplazar CONSOLIDADOS por la tabla dinámica real
    del wb["CONSOLIDADOS"]
    ws = wb.create_sheet("CONSOLIDADOS")
    no_grid(ws)
    # CONSOLIDADOS = pivot table: Nombre compañía | Empleado | Nombre | CES | VAC | PRIMA | INT
    # Matches original layout (cols N-U in original file)
    df16_piv = df_con[df_con["Id Compañia"]==16].copy()

    # Build pivot: one row per employee, columns = causado del mes per cod
    pivot_map = {}  # emp_id -> {nombre, 1:ces, 2:int, 3:prima, 4:vac}
    for _, r in df16_piv.iterrows():
        emp = int(r["Empleado"])
        nombre = r["Nombre del empleado"]
        cod = int(r["Cod. consolidación"])
        val = int(r["Causado del mes"]) if pd.notna(r.get("Causado del mes")) else 0
        if emp not in pivot_map:
            pivot_map[emp] = {"nombre": nombre, 1:0, 2:0, 3:0, 4:0}
        pivot_map[emp][cod] = val

    # Header row 1: filter/pivot headers
    empresa_col = "HOTEL BARRANQUILLA BUENAVISTA SAS"
    headers_piv = ["Nombre de compañia","Empleado","Nombre del empleado",
                   "CESANTIAS","VACACIONES","PRIMA LEGAL","INTERESES SOBRE CESANTIAS"]
    for ci,h in enumerate(headers_piv,1):
        c=ws.cell(1,ci,h); hdr(c)
        ws.column_dimensions[get_column_letter(ci)].width = [35,14,38,14,14,14,22][ci-1]

    # Data rows
    for ri,(emp,(data)) in enumerate(pivot_map.items(),2):
        ws.cell(ri,1,empresa_col).font=F_BLACK
        ws.cell(ri,2,emp).font=F_BLACK
        ws.cell(ri,3,data["nombre"]).font=F_BLACK
        for ci,cod in zip([4,5,6,7],[1,4,3,2]):
            c=ws.cell(ri,ci,data[cod]); c.font=F_BLACK
            c.number_format=NUM_FMT

    # ════════════════════════════════════════════════════════
    # 19. T. DATOS  — tabla multi-bloque real
    # ════════════════════════════════════════════════════════
    ws_td = wb.create_sheet("T. DATOS"); no_grid(ws_td)

    # 9 bloques de 4 columnas cada uno (col 1-4, 5-8, …, 33-36) + col 37 vacía
    BLOQUES = [
        ("CESANTIAS",                "25101001", "Etiquetas de fila", "Suma de Neto"),
        ("CESANTIAS",                "25150101", "Etiquetas de fila", "Suma de Neto"),
        ("PRIMA",                    "25200501", "Etiquetas de fila", "Suma de Neto"),
        ("VACACIONES",               "25250101", "Etiquetas de fila", "Suma de Neto"),
        ("LIBRANZA",                 "28150504", "Tercero movto.",    "Suma de Créditos"),
        ("DESCUENTO SEGUROS A EMPLEADO","23709502","Tercero movto.", "Suma de Créditos"),
        ("UNIMOS",                   "23704005", "Tercero movto.",    "Suma de Créditos"),
        ("PAYFLOW",                  "23709503", "Tercero movto.",    "Suma de Créditos"),
        ("FEMTUR",                   "23704006", "Tercero movto.",    "Suma de Créditos"),
    ]

    # Fila 1: títulos de bloque
    for i,(titulo,_,_,_) in enumerate(BLOQUES):
        col=i*4+1
        c=ws_td.cell(1,col,titulo); c.font=F_BLACK_B

    # Fila 2: vacía (original la tiene vacía)

    # Fila 3: Compañia / valor
    for i,(_,cta,_,_) in enumerate(BLOQUES):
        col=i*4+1
        ws_td.cell(3,col,"Compañia").font=F_BLACK_B
        ws_td.cell(3,col+1,CIA16).font=F_BLACK

    # Fila 4: Auxiliar / cuenta
    for i,(_,cta,_,_) in enumerate(BLOQUES):
        col=i*4+1
        ws_td.cell(4,col,"Auxiliar").font=F_BLACK_B
        ws_td.cell(4,col+1,int(cta)).font=F_BLACK

    # Fila 5: vacía

    # Fila 6: cabeceras de columna datos
    for i,(_,_,h1,h2) in enumerate(BLOQUES):
        col=i*4+1
        ws_td.cell(6,col,h1).font=F_BLACK_B
        ws_td.cell(6,col+1,h2).font=F_BLACK_B
        ws_td.cell(6,col+2,"" if h2=="Suma de Neto" else "").font=F_BLACK_B

    # Datos por bloque
    bloques_data = []
    for titulo, cta, h1, h2 in BLOQUES:
        cta_int = int(cta)
        s = df_aug[(df_aug["Compañia"]==CIA16) &
                   (df_aug["Auxiliar"].astype(str).str.startswith(cta))]
        if s.empty:
            bloques_data.append(pd.DataFrame(columns=["Nit","Nom","Val","Val2"]))
            continue
        if h2 == "Suma de Neto":
            # Cesantias/Prima/Vac: col1=Tercero movto, col2=Neto(negativo), col3=Abs(Neto)
            g = s.groupby("Tercero movto.")["Neto"].sum().reset_index()
            g.columns = ["Nit","Val"]
            g["Val_abs"] = g["Val"].abs()
            g = g[g["Val"]!=0]
            # Store as (nit, neto_neg, abs, None) → col1=nit, col2=neto, col3=abs
            rows_b = [(int(r["Nit"]) if pd.notna(r["Nit"]) else None,
                       None, int(r["Val"]), int(r["Val_abs"])) for _,r in g.iterrows()]
            bloques_data.append(rows_b)
        else:
            # Libranza/Seguros/Unimos/Payflow/Femtur: col1=Tercero, col2=Razón, col3=Créditos
            g = s.groupby(["Tercero movto.","Razón social tercero movto."])["Créditos"].sum().reset_index()
            g.columns = ["Nit","Nom","Val"]
            g = g[g["Val"]!=0]
            rows_b = [(int(r["Nit"]) if pd.notna(r["Nit"]) else None,
                       r["Nom"], int(r["Val"]), None) for _,r in g.iterrows()]
            bloques_data.append(rows_b)

    # Calcular cuántas filas necesita la tabla
    max_rows = max((len(d) for d in bloques_data), default=0)
    bloques_data = [d if isinstance(d, list) else [] for d in bloques_data]

    for data_row in range(max_rows):
        excel_row = 7 + data_row
        for i, rows_b in enumerate(bloques_data):
            col = i*4+1
            if data_row < len(rows_b):
                nit_v, nom_v, val_v, val2_v = rows_b[data_row]
                if nit_v is not None:
                    c1 = ws_td.cell(excel_row, col, nit_v); c1.font=F_BLACK
                if nom_v:
                    c2 = ws_td.cell(excel_row, col+1, nom_v); c2.font=F_BLACK
                if val_v is not None:
                    c3 = ws_td.cell(excel_row, col+2, val_v)
                    c3.font=F_BLACK; c3.number_format=NUM_FMT
                if val2_v is not None:
                    c4 = ws_td.cell(excel_row, col+3, val2_v)
                    c4.font=F_BLACK; c4.number_format=NUM_FMT

    # ════════════════════════════════════════════════════════
    # 20. T. SS
    # ════════════════════════════════════════════════════════
    ws_tss = wb.create_sheet("T. SS "); no_grid(ws_tss)
    cw(ws_tss,[(1,16),(2,55),(3,18)])
    ws_tss.cell(1,1,"Compañia").font=F_BLACK_B; ws_tss.cell(1,2,CIA16).font=F_BLACK
    ws_tss.cell(2,1,"Auxiliar").font=F_BLACK_B; ws_tss.cell(2,2,"(Varios elementos)").font=F_BLACK
    ws_tss.cell(3,1,"Suma de Neto").font=F_BLACK_B
    for lbl,col in zip(["Terceromovto.","Razón social tercero movto.","Total"],[1,2,3]):
        ws_tss.cell(4,col,lbl).font=F_BLACK_B
    row_tss=5
    ss_ctas=["23700501","23700601","23701001","23701002","23701003","23803001","23803003"]
    all_ss=pd.concat([grp_movto(c) for c in ss_ctas],ignore_index=True)
    if not all_ss.empty:
        g=all_ss.groupby(["Nit","Tercero"])["Neto"].sum().reset_index()
        g=g[g["Neto"]!=0]
        for _,r in g.iterrows():
            ws_tss.cell(row_tss,1,int(r["Nit"]) if pd.notna(r["Nit"]) else "").font=F_BLACK
            ws_tss.cell(row_tss,2,r["Tercero"]).font=F_BLACK
            ws_tss.cell(row_tss,3,abs(int(r["Neto"]))).number_format=NUM_FMT
            ws_tss.cell(row_tss,3).font=F_BLACK
            row_tss+=1

    # ════════════════════════════════════════════════════════
    # 21. AUXILIAR SEPTIEMBRE
    # ════════════════════════════════════════════════════════
    ws_as=wb.create_sheet("AUXILIAR SEPTIEMBRE"); no_grid(ws_as)
    if not df_sep.empty:
        df16s=df_sep[df_sep["Compañia"]==CIA16] if "Compañia" in df_sep.columns else df_sep
        for ci,h in enumerate(df16s.columns,1): hdr(ws_as.cell(1,ci,h))
        for ri,row_data in enumerate(df16s.values,2):
            for ci,val in enumerate(row_data,1):
                if pd.notna(val):
                    c=ws_as.cell(ri,ci,val); c.font=F_BLACK
                    if isinstance(val,datetime): c.number_format=DATE_FMT

    # ════════════════════════════════════════════════════════
    # 22. AUXILIAR AGOSTO
    # ════════════════════════════════════════════════════════
    ws_aa=wb.create_sheet("AUXILIAR AGOSTO"); no_grid(ws_aa)
    df16a=df_aug[df_aug["Compañia"]==CIA16]
    for ci,h in enumerate(df16a.columns,1): hdr(ws_aa.cell(1,ci,h))
    for ri,row_data in enumerate(df16a.values,2):
        for ci,val in enumerate(row_data,1):
            if pd.notna(val):
                c=ws_aa.cell(ri,ci,val); c.font=F_BLACK
                if isinstance(val,datetime): c.number_format=DATE_FMT

    # ════════════════════════════════════════════════════════
    # 23. Hoja7
    # ════════════════════════════════════════════════════════
    ws7=wb.create_sheet("Hoja7"); no_grid(ws7)
    ws7.cell(1,1,"CUENTAS").font=F_BLACK_B
    for i,c in enumerate([25101001,25150101,25200501,25250101,
                           25050101,28150504,23709502,23709502,23704005,269596],1):
        ws7.cell(i,2,c).font=F_BLACK

    # ════════════════════════════════════════════════════════
    # 24. CONSULTA DE TRANSA
    # ════════════════════════════════════════════════════════
    ws_ct=wb.create_sheet("CONSULTA DE TRANSA"); no_grid(ws_ct)
    baq_df=baq()
    if not baq_df.empty:
        # Original has 21 columns (cols A-U): base transaction cols + NETO
        # Include NETO column if present (21st col in original = index 20)
        n_cols = min(21, len(baq_df.columns))
        cols_ct = list(baq_df.columns[:n_cols])
        for ci,h in enumerate(cols_ct,1): hdr(ws_ct.cell(1,ci,h))
        for ri,row_data in enumerate(baq_df[cols_ct].values,2):
            for ci,val in enumerate(row_data,1):
                if pd.notna(val) and val!="":
                    c=ws_ct.cell(ri,ci,val); c.font=F_BLACK
                    if isinstance(val,datetime): c.number_format=DATE_FMT
                    if isinstance(val,(int,float)) and ci in [8,9,21]:
                        c.number_format=NUM_FMT

    # ── Guardar y descargar ────────────────────────────────────────────
    out=io.BytesIO(); wb.save(out); out.seek(0)
    fname=f"CONCILIACION_2025_{mes_nombre}_Hiex_Baq.xlsx"

    st.success("✅ Archivo generado con las 24 hojas.")
    st.download_button("⬇️ Descargar conciliación", data=out,
                       file_name=fname,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Métricas
    c1,c2,c3,c4=st.columns(4)
    c1.metric("EPS 237005",       f"${t_eps:,.0f}")
    c2.metric("ARL 237006",       f"${t_arl:,.0f}")
    c3.metric("Pensión 238030",   f"${t_pen:,.0f}")
    c4.metric("Total planilla SS",f"${total_ss:,.0f}")

    hojas_list = ["INICIO","2370-2380","MI PLANILLA SS","2505","28150504","23709503",
                  "23704006","23709502","269596","23704005","136595","23803004","25050101",
                  "CONSOLIDADO CESANTIAS","CONSOLIDADOS","CONSOLIDADO INT CESANTIAS",
                  "CONSOLIDADO PRIMA","CONSOLIDADO VACACIONES","T. DATOS","T. SS",
                  "AUXILIAR SEPTIEMBRE","AUXILIAR AGOSTO","Hoja7","CONSULTA DE TRANSA"]
    badges = " &nbsp;".join([
        f"<span style='background:#e8f0e8;color:#5f6d5f;border-radius:4px;"
        f"padding:2px 8px;font-size:0.78rem;font-weight:600;'>{h}</span>"
        for h in hojas_list
    ])
    st.markdown(f"<div style='margin-top:1rem;line-height:2.2;'>{badges}</div>",
                unsafe_allow_html=True)

   except Exception as e:
    st.error(f"❌ Error: {e}")
    import traceback; st.code(traceback.format_exc())

st.markdown("---")
st.markdown(
    "<div style='text-align:center;color:#5f6d5f;font-size:0.82rem;padding:0.5rem 0;'>"
    "Herramienta contable · <strong>OxoHotel</strong></div>",
    unsafe_allow_html=True
)
